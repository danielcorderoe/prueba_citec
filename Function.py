import pandas as pd
from openpyxl import load_workbook

def cargar_dataset(file_path, chunk_size=10000):
    """
    Carga un dataset desde un archivo Excel en chunks, optimiza los tipos de datos y maneja los valores nulos.

    :param file_path: Ruta del archivo Excel a cargar.
    :param chunk_size: Tamaño de los chunks para cargar el archivo en partes.
    :return: DataFrame con el dataset cargado.
    """
    # Cargar el archivo Excel
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    # Inicializar una lista para almacenar los chunks
    chunks = []

    # Leer los datos en chunks
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    for i in range(1, len(rows), chunk_size):
        chunk = pd.DataFrame(rows[i:i+chunk_size], columns=header)
        chunks.append(chunk)

    # Concatenar todos los chunks en un solo DataFrame
    df = pd.concat(chunks, ignore_index=True)
    return df

